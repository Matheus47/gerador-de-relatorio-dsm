# relatorio_leads_mql.py
# ============================================================
# Sistema completo para análise de Leads/MQL (RD Station / CRM)
# Gera arquivo Excel com abas: Consolidado, Semanais (por mês),
# Auditoria, Validação e Sumário (com gráficos).
# ============================================================

import argparse
import datetime as dt
import re
from typing import Dict, List, Tuple, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter

# ============================================================
# ---------------------- CONSTANTES --------------------------
# ============================================================

# Ordem oficial de exibição no Consolidado e nas Semanais
OFFICIAL_TAGS = [
    "Leads FarmTell Milk",
    "Leads FarmTell Beef",
    "Leads FarmTell Beef Smart",
    "Leads FarmTell Views",
    "Leads PPD Corte",
    "Leads PPD Leite",
    "Leads Consultoria Online Corte",
    "Leads Consultoria Online Leite",
    "Leads Corte",
    "Leads Leite",
    "Leads FarmTell Mills",
    "Sem Tag de Produto",
    "Ft New Beef",
]

# Ordem de prioridade para identificar a Primary Tag com base
# na linha da 1ª conversão (normalizada)
PRIMARY_PRIORITY_NORMALIZED = [
    "leads farmtell milk",
    "leads farmtell beef",
    "leads farmtell beef smart",
    "leads farmtell views",
    "leads ppd corte",
    "leads ppd leite",
    "leads consultoria online corte",
    "leads consultoria online leite",
    "leads farmtell mills",
    "ft new beef",
]

# Mapeamento de nomes de meses (PT) para número
MESES_PT = {
    1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril", 5: "Maio", 6: "Junho",
    7: "Julho", 8: "Agosto", 9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro"
}

MESES_ABREV_PT = {
    "jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
    "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12
}

MESES_NOME_PT = {
    "janeiro": 1, "fevereiro": 2, "marco": 3, "março": 3, "abril": 4, "maio": 5, "junho": 6,
    "julho": 7, "agosto": 8, "setembro": 9, "outubro": 10, "novembro": 11, "dezembro": 12
}

# ============================================================
# ----------------------- UTILITÁRIOS ------------------------
# ============================================================

def normalize_text(s: str) -> str:
    if pd.isna(s):
        return ""
    s = str(s).lower()
    s = re.sub(r"[áàâã]", "a", s)
    s = re.sub(r"[éèê]", "e", s)
    s = re.sub(r"[íìî]", "i", s)
    s = re.sub(r"[óòôõ]", "o", s)
    s = re.sub(r"[úùû]", "u", s)
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def normalize_phone(p) -> str:
    if pd.isna(p):
        return ""
    return re.sub(r"\D", "", str(p))


def fuzzy_match(columns: List[str], target: str) -> Optional[str]:
    target_norm = normalize_text(target)
    for col in columns:
        if normalize_text(col) == target_norm:
            return col
    for col in columns:
        if target_norm in normalize_text(col):
            return col
    return None


def month_name_pt(year: int, month: int) -> str:
    return MESES_PT.get(month, dt.date(year, month, 1).strftime("%B"))

# ============================================================
# --------------------- CARREGAR & MAPEAR --------------------
# ============================================================

def load_base(path: str) -> pd.DataFrame:
    df = pd.read_excel(path)
    df.columns = [str(c) for c in df.columns]
    return df


def map_columns(df: pd.DataFrame) -> Dict[str, Optional[str]]:
    cols = df.columns
    return {
        "primeira_conv": fuzzy_match(cols, "data da primeira conversao"),
        "ultima_opp": fuzzy_match(cols, "data da ultima oportunidade"),
        "origem": fuzzy_match(cols, "origem da primeira conversao"),
        "tags": fuzzy_match(cols, "tags"),
        "email": fuzzy_match(cols, "email"),
        "telefone": fuzzy_match(cols, "telefone"),
        "nome": fuzzy_match(cols, "nome"),
    }


def preprocess(df: pd.DataFrame, cols: Dict[str, Optional[str]]) -> pd.DataFrame:
    # Datas
    df["primeira_conv"] = pd.to_datetime(df[cols["primeira_conv"]], errors="coerce") if cols.get("primeira_conv") else pd.NaT
    df["ultima_opp"] = pd.to_datetime(df[cols["ultima_opp"]], errors="coerce") if cols.get("ultima_opp") else pd.NaT

    # Identificadores
    df["email_clean"] = df[cols["email"]].fillna("").astype(str).str.lower() if cols.get("email") else ""
    df["tel_clean"] = df[cols["telefone"]].apply(normalize_phone) if cols.get("telefone") else ""
    df["nome_clean"] = df[cols["nome"]].fillna("").astype(str) if cols.get("nome") else ""

    def make_id(row):
        if row.get("email_clean", "") != "":
            return row["email_clean"]
        if row.get("tel_clean", "") != "":
            return row["tel_clean"]
        if row.get("nome_clean", "") != "":
            return row["nome_clean"]
        return None

    df["lead_id"] = df.apply(make_id, axis=1)

    # Mantém todas as linhas neste estágio — descartes formais são anotados em Auditoria
    return df

# ============================================================
# ----------------- TAGS (PRIMARY e CL_Hist) ----------------
# ============================================================

def canonicalize_primary_label(norm_label: str) -> str:
    """Converte label normalizado para rótulo canônico da UI."""
    mapping = {
        "leads farmtell milk": "Leads FarmTell Milk",
        "leads farmtell beef": "Leads FarmTell Beef",
        "leads farmtell beef smart": "Leads FarmTell Beef Smart",
        "leads farmtell views": "Leads FarmTell Views",
        "leads ppd corte": "Leads PPD Corte",
        "leads ppd leite": "Leads PPD Leite",
        "leads consultoria online corte": "Leads Consultoria Online Corte",
        "leads consultoria online leite": "Leads Consultoria Online Leite",
        "leads farmtell mills": "Leads FarmTell Mills",
        "ft new beef": "Ft New Beef",
    }
    return mapping.get(norm_label, "Sem Tag de Produto")


def split_tags_cell(cell_val: str) -> List[str]:
    if pd.isna(cell_val):
        return []
    items = [t.strip() for t in str(cell_val).split(",")]
    return [t for t in items if t]


def collect_primary_tags_from_first_conversion(df: pd.DataFrame, cols: Dict[str, Optional[str]]) -> Dict[str, List[str]]:
    """Coleta as tags APENAS da linha da 1ª conversão (por lead)."""
    tag_map: Dict[str, List[str]] = {}
    for lid, g in df.groupby("lead_id"):
        g_valid = g.dropna(subset=["primeira_conv"]).sort_values("primeira_conv")
        if g_valid.empty:
            tag_map[lid] = []
            continue
        row = g_valid.iloc[0]
        tags_cell = row[cols["tags"]] if cols.get("tags") else ""
        tag_map[lid] = split_tags_cell(tags_cell)
    return tag_map


def collect_all_tags(df: pd.DataFrame, cols: Dict[str, Optional[str]]) -> Dict[str, List[str]]:
    """Coleta todas as tags (histórico completo) por lead (para CL_Hist)."""
    tag_map: Dict[str, List[str]] = {}
    for lid, g in df.groupby("lead_id"):
        tags_raw: List[str] = []
        if cols.get("tags"):
            for val in g[cols["tags"]].fillna(""):
                tags_raw.extend(split_tags_cell(val))
        tag_map[lid] = tags_raw
    return tag_map


def classify_primary_from_list(tag_list: List[str]) -> str:
    tags_norm = [normalize_text(t) for t in tag_list]
    # Caso especial: qualquer tag que COMEÇA com "leads farmtell views"
    for t in tags_norm:
        if t.startswith("leads farmtell views"):
            return "Leads FarmTell Views"
    # Ordem prioritária
    for pri in PRIMARY_PRIORITY_NORMALIZED:
        if pri in tags_norm:
            return canonicalize_primary_label(pri)
    return "Sem Tag de Produto"


def cl_historico(all_tags: List[str]) -> str:
    tags_norm = [normalize_text(t) for t in all_tags]
    filtered = [t for t in tags_norm if not t.startswith("leads farmtell views")]
    for t in filtered:
        if "leads leite" in t or "lead leite" in t:
            return "Leads Leite"
    for t in filtered:
        if "leads corte" in t or "lead corte" in t:
            return "Leads Corte"
    return ""

# ============================================================
# ------------------------- MQL MAP --------------------------
# ============================================================

def get_mql_map(df: pd.DataFrame) -> Dict[str, bool]:
    mql_map: Dict[str, bool] = {}
    for lid, g in df.groupby("lead_id"):
        has_opp = g["ultima_opp"].dropna().shape[0] > 0
        mql_map[lid] = has_opp
    return mql_map

# ============================================================
# ----------------- INTERVALOS MENSAL/SEMANAL ----------------
# ============================================================

def first_day_of_month(d: dt.date) -> dt.date:
    return dt.date(d.year, d.month, 1)


def next_month(d: dt.date) -> dt.date:
    return dt.date(d.year + (1 if d.month == 12 else 0), 1 if d.month == 12 else d.month + 1, 1)


def build_month_range(start_month: dt.date, end_month: dt.date) -> List[dt.date]:
    months: List[dt.date] = []
    cur = first_day_of_month(start_month)
    last = first_day_of_month(end_month)
    while cur <= last:
        months.append(cur)
        cur = next_month(cur)
    return months


def build_weeks_dom_sab(start_date: dt.date, end_date: dt.date) -> List[Tuple[dt.date, dt.date]]:
    delta = start_date.weekday()  # Mon=0..Sun=6
    if delta != 6:
        start = start_date - dt.timedelta(days=delta + 1)
    else:
        start = start_date
    weeks: List[Tuple[dt.date, dt.date]] = []
    cur = start
    while cur <= end_date:
        w_start = cur
        w_end = cur + dt.timedelta(days=6)
        weeks.append((w_start, w_end))
        cur = cur + dt.timedelta(days=7)
    return weeks


def intersect_week_with_month(week: Tuple[dt.date, dt.date], m0: dt.date, m1: dt.date) -> Optional[Tuple[dt.date, dt.date]]:
    ws, we = week
    s = max(ws, m0)
    e = min(we, m1)
    if s > e:
        return None
    return (s, e)

# ============================================================
# ------------------------ CONSOLIDADO -----------------------
# ============================================================

def gerar_consolidado(
    df: pd.DataFrame,
    primary_map: Dict[str, str],
    cl_map: Dict[str, str],
    months: List[dt.date],
) -> pd.DataFrame:
    registros: List[Dict[str, object]] = []

    # Mapeia lead -> data da última oportunidade (a mais recente)
    lead_last_opp: Dict[str, Optional[dt.date]] = {}
    for lid, g in df.groupby("lead_id"):
        opps = g["ultima_opp"].dropna().sort_values()
        lead_last_opp[lid] = opps.iloc[-1].date() if not opps.empty else None

    for m0 in months:
        m1 = next_month(m0)
        # Conjunto de leads do mês (pela primeira conversão)
        mask = (df["primeira_conv"] >= pd.Timestamp(m0)) & (df["primeira_conv"] < pd.Timestamp(m1))
        leads_month = set(df.loc[mask, "lead_id"].dropna().unique())

        cnt_leads = {tag: 0 for tag in OFFICIAL_TAGS}
        cnt_mqls = {tag: 0 for tag in OFFICIAL_TAGS}

        # Leads por Primary e CL
        for lid in leads_month:
            ptag = primary_map.get(lid, "Sem Tag de Produto")
            if ptag not in OFFICIAL_TAGS:
                ptag = "Sem Tag de Produto"
            cnt_leads[ptag] += 1
            cl = cl_map.get(lid, "")
            if cl in ("Leads Corte", "Leads Leite"):
                cnt_leads[cl] += 1

        # MQLs do mês: todos os leads cuja última opp cai no mês (independente da 1ª conv)
        for lid, d_opp in lead_last_opp.items():
            if d_opp is None:
                continue
            if not (m0 <= d_opp < m1):
                continue
            ptag = primary_map.get(lid, "Sem Tag de Produto")
            if ptag not in OFFICIAL_TAGS:
                ptag = "Sem Tag de Produto"
            cnt_mqls[ptag] += 1
            cl = cl_map.get(lid, "")
            if cl in ("Leads Corte", "Leads Leite"):
                cnt_mqls[cl] += 1

        for tag in OFFICIAL_TAGS:
            registros.append({
                "Mes": f"{m0.year:04d}-{m0.month:02d}",
                "Tag": tag,
                "Leads": int(cnt_leads[tag]),
                "MQLs": int(cnt_mqls[tag]),
            })

    consolidado = pd.DataFrame(registros)
    consolidado["Tag"] = pd.Categorical(consolidado["Tag"], categories=OFFICIAL_TAGS, ordered=True)
    consolidado.sort_values(["Mes", "Tag"], inplace=True)
    consolidado.reset_index(drop=True, inplace=True)
    return consolidado

# ============================================================
# -------------------------- SEMANAIS ------------------------
# ============================================================

def assign_date_to_week(d: Optional[pd.Timestamp], month_weeks: List[Tuple[dt.date, dt.date]]) -> Optional[int]:
    if d is None or pd.isna(d):
        return None
    the_date = d.date()
    for i, (ws, we) in enumerate(month_weeks):
        if ws <= the_date <= we:
            return i
    return None


def gerar_semanais_por_mes(
    df: pd.DataFrame,
    primary_map: Dict[str, str],
    cl_map: Dict[str, str],
    weeks_global: List[Tuple[dt.date, dt.date]],
    month_date: dt.date,
) -> pd.DataFrame:
    m0 = dt.date(month_date.year, month_date.month, 1)
    m1 = next_month(m0)
    m_last = m1 - dt.timedelta(days=1)

    month_weeks: List[Tuple[dt.date, dt.date]] = []
    for w in weeks_global:
        inter = intersect_week_with_month(w, m0, m_last)
        if inter:
            month_weeks.append(inter)

    week_labels = [f"{s.strftime('%d/%m/%Y')} a {e.strftime('%d/%m/%Y')}" for (s, e) in month_weeks]

    # Conjunto de leads do mês para LEADS
    mask_leads = (df["primeira_conv"] >= pd.Timestamp(m0)) & (df["primeira_conv"] < pd.Timestamp(m1))
    leads_month_ids = set(df.loc[mask_leads, "lead_id"].dropna().unique())

    # Última opp por lead (para MQLs, sem limitar 1ª conv)
    lead_last_opp: Dict[str, Optional[dt.date]] = {}
    for lid, g in df.groupby("lead_id"):
        opps = g["ultima_opp"].dropna().sort_values()
        lead_last_opp[lid] = opps.iloc[-1].date() if not opps.empty else None

    rows: List[List] = []
    for tag in OFFICIAL_TAGS:
        rows.append(["SEMANA", *week_labels, "Mês Fechado"])

        # Leads
        linha_leads = ["Leads"]
        for i, _ in enumerate(month_weeks):
            count = 0
            for lid in leads_month_ids:
                ptag = primary_map.get(lid, "Sem Tag de Produto")
                cl = cl_map.get(lid, "")
                pertence = (tag in ("Leads Corte", "Leads Leite") and cl == tag) or \
                           (tag not in ("Leads Corte", "Leads Leite") and ptag == tag)
                if not pertence:
                    continue
                d_first_series = df.loc[df["lead_id"] == lid, "primeira_conv"].dropna()
                if d_first_series.empty:
                    continue
                wk = assign_date_to_week(d_first_series.iloc[0], month_weeks)
                if wk == i:
                    count += 1
            linha_leads.append(count)
        linha_leads.append(None)
        rows.append(linha_leads)

        # MQLs
        linha_mqls = ["MQLs"]
        for i, (ws, we) in enumerate(month_weeks):
            count = 0
            for lid, d_opp in lead_last_opp.items():
                if d_opp is None:
                    continue
                if not (m0 <= d_opp < m1):
                    continue
                ptag = primary_map.get(lid, "Sem Tag de Produto")
                cl = cl_map.get(lid, "")
                pertence = (tag in ("Leads Corte", "Leads Leite") and cl == tag) or \
                           (tag not in ("Leads Corte", "Leads Leite") and ptag == tag)
                if not pertence:
                    continue
                if ws <= d_opp <= we:
                    count += 1
            linha_mqls.append(count)
        linha_mqls.append(None)
        rows.append(linha_mqls)

        rows.append([""] * (len(week_labels) + 2))

    max_cols = max(len(r) for r in rows) if rows else 0
    rows = [r + [""] * (max_cols - len(r)) for r in rows]
    return pd.DataFrame(rows)


def gerar_semanais(
    df: pd.DataFrame,
    primary_map: Dict[str, str],
    cl_map: Dict[str, str],
    intervalo_mensal: Tuple[dt.date, dt.date],
    intervalo_semanal: Tuple[dt.date, dt.date],
) -> Dict[str, pd.DataFrame]:
    start_m, end_m = intervalo_mensal
    months = build_month_range(start_m, end_m)
    weeks_global = build_weeks_dom_sab(*intervalo_semanal)
    abas: Dict[str, pd.DataFrame] = {}
    for m in months:
        nome = month_name_pt(m.year, m.month)
        abas[nome] = gerar_semanais_por_mes(df, primary_map, cl_map, weeks_global, m)
    return abas

# ============================================================
# -------------------------- AUDITORIA -----------------------
# ============================================================

def gerar_auditoria(
    df: pd.DataFrame,
    cols: Dict[str, Optional[str]],
    primary_map: Dict[str, str],
    cl_map: Dict[str, str],
    intervalo_mensal: Tuple[dt.date, dt.date],
) -> Tuple[pd.DataFrame, Dict[str, int]]:
    start_m, end_m = intervalo_mensal
    m0 = first_day_of_month(start_m)
    m_end_plus = next_month(first_day_of_month(end_m))

    mask = (df["primeira_conv"] >= pd.Timestamp(m0)) & (df["primeira_conv"] < pd.Timestamp(m_end_plus))
    df_recorte = df.loc[mask].copy()

    registros: List[Dict[str, object]] = []
    descartes = {
        "linhas_descartadas_sem_primeira_conversao": 0,
        "linhas_descartadas_sem_lead_id": 0,
    }

    for lid, g in df_recorte.groupby("lead_id"):
        g_valid = g.dropna(subset=["primeira_conv"]).sort_values("primeira_conv")
        if g_valid.empty:
            descartes["linhas_descartadas_sem_primeira_conversao"] += len(g)
            continue
        row = g_valid.iloc[0]
        origem = row[cols["origem"]] if cols.get("origem") else ""
        tags_brutas = row[cols["tags"]] if cols.get("tags") else ""
        if pd.isna(lid) or lid == "":
            descartes["linhas_descartadas_sem_lead_id"] += len(g)
            continue
        registros.append({
            "lead_id": lid,
            "Data da primeira conversao": row["primeira_conv"],
            "Primary Tag": primary_map.get(lid, "Sem Tag de Produto"),
            "CL_Hist": cl_map.get(lid, ""),
            "MQL?": "Sim" if not g["ultima_opp"].dropna().empty else "Não",
            "Origem da primeira conversao": origem,
            "Tags originais": tags_brutas,
        })

    aud_df = pd.DataFrame(registros)
    aud_df.sort_values(by=["Primary Tag", "CL_Hist", "Data da primeira conversao"], inplace=True)
    aud_df.reset_index(drop=True, inplace=True)
    return aud_df, descartes

# ============================================================
# -------------------------- VALIDAÇÃO -----------------------
# ============================================================

def somar_semanais_mes(df_semanais_mes: pd.DataFrame) -> Dict[str, Dict[str, int]]:
    resultados: Dict[str, Dict[str, int]] = {}
    linhas = df_semanais_mes.values.tolist()
    idx = 0
    tag_idx = 0
    while idx < len(linhas) and tag_idx < len(OFFICIAL_TAGS):
        tag = OFFICIAL_TAGS[tag_idx]
        linha_leads = linhas[idx + 1]
        linha_mqls = linhas[idx + 2]
        vals_leads = [x for x in linha_leads[1:-1] if isinstance(x, (int, float))]
        vals_mqls = [x for x in linha_mqls[1:-1] if isinstance(x, (int, float))]
        resultados[tag] = {
            "sem_leads": int(sum(vals_leads)) if vals_leads else 0,
            "sem_mqls": int(sum(vals_mqls)) if vals_mqls else 0,
        }
        idx += 4
        tag_idx += 1
    return resultados


def gerar_validacao(consolidado_df: pd.DataFrame, abas_semanais: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    registros: List[Dict[str, object]] = []
    for mes in consolidado_df["Mes"].unique():
        ano = int(mes.split("-")[0]); mes_n = int(mes.split("-")[1])
        nome_aba = month_name_pt(ano, mes_n)
        if nome_aba not in abas_semanais:
            continue
        df_sem = abas_semanais[nome_aba]
        resumo_sem = somar_semanais_mes(df_sem)
        df_mes = consolidado_df[consolidado_df["Mes"] == mes]
        for _, row in df_mes.iterrows():
            tag = row["Tag"]
            cons_leads = int(row["Leads"])
            cons_mqls = int(row["MQLs"])
            sem_leads = resumo_sem.get(tag, {}).get("sem_leads", 0)
            sem_mqls = resumo_sem.get(tag, {}).get("sem_mqls", 0)
            registros.append({
                "Mes": mes,
                "Tag": tag,
                "Consolidado_Leads": cons_leads,
                "Semanal_Leads": sem_leads,
                "Status_Leads": "OK" if cons_leads == sem_leads else "ALERTA",
                "Consolidado_MQLs": cons_mqls,
                "Semanal_MQLs": sem_mqls,
                "Status_MQLs": "OK" if cons_mqls == sem_mqls else "ALERTA",
            })
    val_df = pd.DataFrame(registros)
    val_df["Tag"] = pd.Categorical(val_df["Tag"], categories=OFFICIAL_TAGS, ordered=True)
    val_df.sort_values(["Mes", "Tag"], inplace=True)
    val_df.reset_index(drop=True, inplace=True)
    return val_df

# ============================================================
# --------------------------- SUMÁRIO ------------------------
# ============================================================

def montar_tabela_leads(consolidado_df: pd.DataFrame) -> pd.DataFrame:
    meses = sorted(consolidado_df["Mes"].unique())
    data = {"Tag": OFFICIAL_TAGS}
    for mes in meses:
        vals = []
        df_mes = consolidado_df[consolidado_df["Mes"] == mes]
        for tag in OFFICIAL_TAGS:
            r = df_mes[df_mes["Tag"] == tag]
            vals.append(int(r["Leads"].iloc[0]) if not r.empty else 0)
        data[mes] = vals
    return pd.DataFrame(data)


def montar_tabela_mqls(consolidado_df: pd.DataFrame) -> pd.DataFrame:
    meses = sorted(consolidado_df["Mes"].unique())
    data = {"Tag": OFFICIAL_TAGS}
    for mes in meses:
        vals = []
        df_mes = consolidado_df[consolidado_df["Mes"] == mes]
        for tag in OFFICIAL_TAGS:
            r = df_mes[df_mes["Tag"] == tag]
            vals.append(int(r["MQLs"].iloc[0]) if not r.empty else 0)
        data[mes] = vals
    return pd.DataFrame(data)


def montar_tabela_conversao(tab_leads: pd.DataFrame, tab_mqls: pd.DataFrame) -> pd.DataFrame:
    meses = [c for c in tab_leads.columns if c != "Tag"]
    data = {"Tag": OFFICIAL_TAGS}
    for mes in meses:
        vals = []
        for tag in OFFICIAL_TAGS:
            L = float(tab_leads.loc[tab_leads["Tag"] == tag, mes].iloc[0])
            M = float(tab_mqls.loc[tab_mqls["Tag"] == tag, mes].iloc[0])
            vals.append((M / L) if L > 0 else 0.0)
        data[mes] = vals
    return pd.DataFrame(data)


def gerar_sumario(consolidado_df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    tab_leads = montar_tabela_leads(consolidado_df)
    tab_mqls = montar_tabela_mqls(consolidado_df)
    tab_conv = montar_tabela_conversao(tab_leads, tab_mqls)
    return {"tabela_leads": tab_leads, "tabela_mqls": tab_mqls, "tabela_conversao": tab_conv}

# ============================================================
# ----------------------- EXCEL WRITER PRO -------------------
# ============================================================

BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
FILL_HEADER = PatternFill("solid", fgColor="D9D9D9")
FILL_BLOCK = PatternFill("solid", fgColor="F2F2F2")


def auto_adjust(ws):
    for col in ws.columns:
        max_len = 0
        column = col[0].column_letter
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
                if len(val) > max_len:
                    max_len = len(val)
            except Exception:
                pass
        ws.column_dimensions[column].width = min(max_len + 2, 60)


def write_df(ws, df: pd.DataFrame, bold_header: bool = True, border: bool = True):
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    if bold_header:
        for c in ws[1]:
            c.font = BOLD
            c.fill = FILL_HEADER
            c.alignment = CENTER
    if border:
        for row in ws.iter_rows():
            for c in row:
                c.border = THIN_BORDER


def escrever_aba_semanal(ws, df_sem: pd.DataFrame):
    for r in dataframe_to_rows(df_sem, index=False, header=False):
        ws.append(r)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        label = row[0].value
        if label == "SEMANA":
            for c in row:
                c.font = BOLD
                c.fill = FILL_BLOCK
                c.alignment = CENTER
        elif label in ("Leads", "MQLs"):
            for c in row:
                c.alignment = CENTER
        for c in row:
            c.border = THIN_BORDER
    if ws.max_row > 0 and ws.max_column > 2:
        last_col = get_column_letter(ws.max_column)
        for r in range(1, ws.max_row + 1):
            if ws[f"A{r}"].value in ("Leads", "MQLs"):
                ws[f"{last_col}{r}"] = f"=SUM({get_column_letter(2)}{r}:{get_column_letter(ws.max_column-1)}{r})"
    ws.freeze_panes = "A2"
    auto_adjust(ws)


def escrever_auditoria(ws, df_aud: pd.DataFrame, descartes: Dict[str, int]):
    write_df(ws, df_aud)
    start = ws.max_row + 2
    ws[f"A{start}"] = "Descartes:"
    ws[f"A{start}"].font = BOLD
    row = start + 1
    for motivo, qtd in descartes.items():
        ws[f"A{row}"] = f"{motivo}"
        ws[f"B{row}"] = int(qtd)
        ws[f"A{row}"].font = BOLD
        row += 1
    ws.freeze_panes = "A2"
    auto_adjust(ws)


def criar_graficos_sumario(ws, lead_table_start: int, tab_leads: pd.DataFrame,
                           mql_table_start: int, tab_mqls: pd.DataFrame,
                           conv_table_start: int, tab_conv: pd.DataFrame):
    # Leads (barras)
    n_rows_leads = len(tab_leads) + 1
    n_cols_leads = len(tab_leads.columns)
    data = Reference(ws, min_col=2, max_col=n_cols_leads, min_row=lead_table_start, max_row=lead_table_start + n_rows_leads - 1)
    cats = Reference(ws, min_col=1, min_row=lead_table_start + 1, max_row=lead_table_start + n_rows_leads - 1)
    ch1 = BarChart(); ch1.title = "Leads por mês (Consolidado)"; ch1.y_axis.title = "Leads"
    ch1.add_data(data, titles_from_data=True); ch1.set_categories(cats)
    ws.add_chart(ch1, f"A{conv_table_start + len(tab_conv) + 6}")

    # MQLs (barras)
    n_rows_mql = len(tab_mqls) + 1
    n_cols_mql = len(tab_mqls.columns)
    data2 = Reference(ws, min_col=2, max_col=n_cols_mql, min_row=mql_table_start, max_row=mql_table_start + n_rows_mql - 1)
    cats2 = Reference(ws, min_col=1, min_row=mql_table_start + 1, max_row=mql_table_start + n_rows_mql - 1)
    ch2 = BarChart(); ch2.title = "MQLs por mês (Consolidado)"; ch2.y_axis.title = "MQLs"
    ch2.add_data(data2, titles_from_data=True); ch2.set_categories(cats2)
    ws.add_chart(ch2, f"I{conv_table_start + len(tab_conv) + 6}")

    # Conversão (%) (linha)
    n_rows_conv = len(tab_conv) + 1
    n_cols_conv = len(tab_conv.columns)
    data3 = Reference(ws, min_col=2, max_col=n_cols_conv, min_row=conv_table_start, max_row=conv_table_start + n_rows_conv - 1)
    cats3 = Reference(ws, min_col=1, min_row=conv_table_start + 1, max_row=conv_table_start + n_rows_conv - 1)
    ch3 = LineChart(); ch3.title = "Conversão MQL/Leads (%) por mês"; ch3.y_axis.title = "Conversão (%)"
    ch3.add_data(data3, titles_from_data=True); ch3.set_categories(cats3); ch3.y_axis.number_format = "0.0%"
    ws.add_chart(ch3, f"A{conv_table_start + len(tab_conv) + 22}")


def gerar_arquivo_final(
    consolidado: pd.DataFrame,
    abas_semanais: Dict[str, pd.DataFrame],
    auditoria_df: pd.DataFrame,
    descartes: Dict[str, int],
    validacao_df: pd.DataFrame,
    sumario_dict: Dict[str, pd.DataFrame],
    nome_arquivo: str,
):
    wb = Workbook(); wb.remove(wb.active)

    ws_con = wb.create_sheet("Consolidado"); write_df(ws_con, consolidado); ws_con.freeze_panes = "A2"; auto_adjust(ws_con)

    for nome, df_sem in abas_semanais.items():
        ws = wb.create_sheet(nome); escrever_aba_semanal(ws, df_sem)

    ws_aud = wb.create_sheet("Auditoria"); escrever_auditoria(ws_aud, auditoria_df, descartes)

    ws_val = wb.create_sheet("Validação"); write_df(ws_val, validacao_df); ws_val.freeze_panes = "A2"
    for r in range(2, ws_val.max_row + 1):
        for col in (5, 8):
            cell = ws_val.cell(row=r, column=col)
            if cell.value == "OK":
                cell.fill = PatternFill("solid", fgColor="C6EFCE")
            elif cell.value == "ALERTA":
                cell.fill = PatternFill("solid", fgColor="FFC7CE")
    auto_adjust(ws_val)

    ws_sum = wb.create_sheet("Sumário")
    row = 1
    ws_sum[f"A{row}"] = "Leads por Tag × Mês"; ws_sum[f"A{row}"].font = BOLD; row += 2
    lead_table_start = row
    for r in dataframe_to_rows(sumario_dict["tabela_leads"], index=False, header=True): ws_sum.append(r)
    for c in next(ws_sum.iter_rows(min_row=lead_table_start, max_row=lead_table_start, min_col=1, max_col=ws_sum.max_column)): c.font = BOLD; c.fill = FILL_HEADER; c.alignment = CENTER
    row = ws_sum.max_row + 2

    ws_sum[f"A{row}"] = "MQLs por Tag × Mês"; ws_sum[f"A{row}"].font = BOLD; row += 2
    mql_table_start = row
    for r in dataframe_to_rows(sumario_dict["tabela_mqls"], index=False, header=True): ws_sum.append(r)
    for c in next(ws_sum.iter_rows(min_row=mql_table_start, max_row=mql_table_start, min_col=1, max_col=ws_sum.max_column)): c.font = BOLD; c.fill = FILL_HEADER; c.alignment = CENTER
    row = ws_sum.max_row + 2

    ws_sum[f"A{row}"] = "Conversão (%) por Tag × Mês"; ws_sum[f"A{row}"].font = BOLD; row += 2
    conv_table_start = row
    for r in dataframe_to_rows(sumario_dict["tabela_conversao"], index=False, header=True): ws_sum.append(r)
    n_cols_conv = len(sumario_dict["tabela_conversao"].columns)
    for rr in range(conv_table_start + 1, conv_table_start + 1 + len(sumario_dict["tabela_conversao"])):
        for cc in range(2, n_cols_conv + 1):
            ws_sum.cell(row=rr, column=cc).number_format = "0.0%"
    for c in next(ws_sum.iter_rows(min_row=conv_table_start, max_row=conv_table_start, min_col=1, max_col=ws_sum.max_column)): c.font = BOLD; c.fill = FILL_HEADER; c.alignment = CENTER

    auto_adjust(ws_sum)

    criar_graficos_sumario(ws_sum, lead_table_start, sumario_dict["tabela_leads"], mql_table_start, sumario_dict["tabela_mqls"], conv_table_start, sumario_dict["tabela_conversao"])

    wb.save(nome_arquivo)

# ============================================================
# ------------------------ PARSERS CLI -----------------------
# ============================================================

def parse_intervalo_mensal(texto: str) -> Tuple[dt.date, dt.date]:
    partes = [p.strip() for p in texto.split(" a ")]
    if len(partes) != 2:
        raise ValueError("INTERVALO MENSAL inválido. Ex.: 'Fev/2026 a Fev/2026'")
    def parse_mes_ano(s: str) -> Tuple[int, int]:
        s_norm = normalize_text(s)
        m = re.match(r"^(\d{1,2})/(\d{4})$", s_norm)
        if m:
            return int(m.group(1)), int(m.group(2))
        m = re.match(r"^([a-z]{3})/(\d{4})$", s_norm)
        if m and m.group(1) in MESES_ABREV_PT:
            return MESES_ABREV_PT[m.group(1)], int(m.group(2))
        m = re.match(r"^([a-zç]+)/(\d{4})$", s_norm)
        if m:
            nome = m.group(1)
            if nome in MESES_NOME_PT:
                return MESES_NOME_PT[nome], int(m.group(2))
        raise ValueError(f"Mês/ano inválido: {s}")
    m1, a1 = parse_mes_ano(partes[0]); m2, a2 = parse_mes_ano(partes[1])
    return dt.date(a1, m1, 1), dt.date(a2, m2, 1)


def parse_intervalo_semanal(texto: str) -> Tuple[dt.date, dt.date]:
    partes = [p.strip() for p in texto.split(" a ")]
    if len(partes) != 2:
        raise ValueError("INTERVALO SEMANAL inválido. Ex.: '01/02/2026 a 28/02/2026'")
    def parse_d(s: str) -> dt.date:
        return dt.datetime.strptime(s, "%d/%m/%Y").date()
    return parse_d(partes[0]), parse_d(partes[1])

# ============================================================
# --------------------------- MAIN ---------------------------
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="Gera relatório de Leads/MQL com abas Consolidado, Semanais, Auditoria, Validação e Sumário.")
    parser.add_argument("--base", required=False, default="base_leads.xlsx", help="Arquivo Excel de entrada (aba com todos os registros).")
    parser.add_argument("--intervalo_mensal", required=False, default="Fev/2026 a Fev/2026", help="Ex.: 'Dez/2025 a Fev/2026'")
    parser.add_argument("--intervalo_semanal", required=False, default="01/02/2026 a 28/02/2026", help="Ex.: '01/12/2025 a 22/02/2026'")
    parser.add_argument("--saida", required=False, default="relatorio_final_leads_mql.xlsx", help="Nome do arquivo de saída (.xlsx)")

    args = parser.parse_args()

    # 1) Carrega base e mapeia
    df_raw = load_base(args.base)
    cols = map_columns(df_raw)

    # 2) Pré-processa (coerção de datas, ID, etc.)
    df = preprocess(df_raw.copy(), cols)

    # 3) Construções de Tag/Origem
    primary_tags_source = collect_primary_tags_from_first_conversion(df, cols)
    primary_map: Dict[str, str] = {lid: classify_primary_from_list(tags) for lid, tags in primary_tags_source.items()}
    all_tags_source = collect_all_tags(df, cols)
    cl_map: Dict[str, str] = {lid: cl_historico(tags) for lid, tags in all_tags_source.items()}

    # 4) Intervalos
    intervalo_mensal = parse_intervalo_mensal(args.intervalo_mensal)
    intervalo_semanal = parse_intervalo_semanal(args.intervalo_semanal)
    months = build_month_range(*intervalo_mensal)

    # 5) Consolidado
    consolidado = gerar_consolidado(df, primary_map, cl_map, months)

    # 6) Semanais (por mês)
    abas_semanais = gerar_semanais(df, primary_map, cl_map, intervalo_mensal, intervalo_semanal)

    # 7) Auditoria (recorte mensal)
    auditoria_df, descartes = gerar_auditoria(df, cols, primary_map, cl_map, intervalo_mensal)

    # 8) Validação (Consolidado x Semanais)
    validacao_df = gerar_validacao(consolidado, abas_semanais)

    # 9) Sumário (tabelas + gráficos no writer)
    sumario_dict = gerar_sumario(consolidado)

    # 10) Excel final
    gerar_arquivo_final(consolidado, abas_semanais, auditoria_df, descartes, validacao_df, sumario_dict, args.saida)

    print(f"Arquivo gerado: {args.saida}")


if __name__ == "__main__":
    main()
